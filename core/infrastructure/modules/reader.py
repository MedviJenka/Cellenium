import json
import openpyxl
from typing import Optional
from PIL import Image
from configparser import ConfigParser
from core.infrastructure.constants.data import *
from core.infrastructure.modules.logger import Logger


log = Logger()


def read_config(key: str, value: str) -> str:
    config = ConfigParser()
    config.read(CONFIG_PATH)
    return config.get(key, value)


def read_json(path: str, value: Optional[str] = None) -> str | dict:
    with open(path, 'r', encoding='utf-8') as json_file:
        if value:
            return json.load(json_file)[value]
        return json.load(json_file)


def write_json(path: str, key: str, value: Optional[any]) -> None:
    data = read_json(path)
    data[key] = value
    with open(path, 'w', encoding='utf-8') as json_file:
        json.dump(data, json_file)


def _read_excel(sheet_name: str, value: str) -> dict[str]:

    workbook = openpyxl.load_workbook(PAGE_BASE)
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows(min_row=2, values_only=True):

        name, locator, element_type, actions, image = row[:5]

        if name == value:
            return {
                'name': name,
                'locator': locator,
                'type': element_type,
                'actions': actions,
                'image': image
            }


def get_name(sheet_name: str, value: str) -> str:
    return _read_excel(sheet_name, value)['name']


def get_locator(sheet_name: str, value: str) -> str:
    return _read_excel(sheet_name, value)['locator']


def get_type(sheet_name: str, value: str) -> str:
    return _read_excel(sheet_name, value)['type']


def get_actions(sheet_name: str, value: str) -> str:
    return _read_excel(sheet_name, value)['actions']


def get_image(sheet_name: str, value: str) -> str:
    return _read_excel(sheet_name, value)['image']


def write_excel(sheet_name: str, value: str) -> None:
    width = 100
    height = 100
    img = Image.open(value)
    img = img.resize((width, height), Image.NEAREST)
    img.save(value)
    workbook = openpyxl.load_workbook(PAGE_BASE)
    sheet = workbook[sheet_name]
    sheet.add_image(img, 'D2')
    workbook.save(fr'{GLOBAL_PATH}\{PAGE_BASE}')


def read_test_case(sheet_name: list[str]) -> list[str]:

    test_case = fr"{GLOBAL_PATH}\{TEST_SUITE}"
    workbook = openpyxl.load_workbook(test_case)
    test_dir = fr"{GLOBAL_PATH}\{TESTS}"
    sheet = []
    for each_sheet_name in sheet_name:
        sheet = workbook[each_sheet_name]
        # print(sheet)
    lists = []
    for row in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
        result = {
            "test": row[0],
            "run": row[1],
        }

        for _, value in result.items():

            if value == '.':
                case = fr'{test_dir}\{sheet.title}\{result["test"]}'
                lists.append(case)

    return lists
