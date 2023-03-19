import os
import openpyxl
from dataclasses import dataclass
from core.infrastructure.constants.data import PROJECT_PATH
from core.infrastructure.modules.executor import Executor
from core.infrastructure.modules.reader import read_config


@dataclass
class RunSuite(Executor):

    sheet_name: list[str]

    def get_sheet_titles(self) -> list[str]:
        test_case = fr"{PROJECT_PATH}\{read_config('path', 'test_case')}"
        workbook = openpyxl.load_workbook(test_case)
        _list = []
        for each_sheet_name in self.sheet_name:
            sheet = workbook[each_sheet_name]
            _list.append(sheet.title)
        return _list

    def read_test_case(self) -> None:
        sheet_title = self.get_sheet_titles()
        report_dir = fr"{PROJECT_PATH}\{read_config('path', 'allure')}"
        test_case = fr"{PROJECT_PATH}\{read_config('path', 'test_case')}"

        # workbook reads the path C:\Users\medvi\OneDrive\Desktop\Cellenium\core\static\utils\test_case.xlsx
        workbook = openpyxl.load_workbook(test_case)
        test_dir = fr"{PROJECT_PATH}\{read_config('path', 'tests')}"

        for each_sheet_name in sheet_title:
            sheet = workbook[each_sheet_name]

            for row in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
                result = {
                    "test": row[0],
                    "run": row[1],
                }

                for _, value in result.items():
                    if value == '.':
                        os.system(fr'pytest {test_dir}\{sheet.title}\{result["test"]} --alluredir={report_dir}')
        os.system(fr'allure serve {report_dir}')

    def execute(self) -> None:
        self.get_sheet_titles()
        self.read_test_case()


if __name__ == '__main__':
    run = RunSuite(['app1', 'google', 'module', 'weathermap'])
    run.execute()


