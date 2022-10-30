import openpyxl
from dataclasses import dataclass
from core.components.config.reader import ConfigReader
import csv


@dataclass
class ExcelReader:

    config = ConfigReader()
    path: str = config.read('path', 'page_base')
    screenshots: str = config.read('path', 'screenshots')

    def _read(self, sheet_name: str, value: str) -> dict[str]:
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[sheet_name]
        cache = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            result = {
                'name': row[0],
                'locator': row[1],
                'type': row[2],
                'image': row[3]
            }
            cache[result['name']] = result

        if cache[value]['name']:
            try:
                return {
                    'name': cache[value]['name'],
                    'locator': cache[value]['locator'],
                    'type': cache[value]['type'],
                    'image': cache[value]['image']
                }
            except Exception:
                raise "no such name in the cell sheet"

    def get_name(self, *args: str) -> str:
        return self._read(*args)['name']

    def get_locator(self, *args: str) -> str:
        return self._read(*args)['locator']

    def get_type(self, *args: str) -> str:
        return self._read(*args)['type']

    def get_image(self, *args: str) -> str:
        return self._read(*args)['image']

    def embed_image_into_cell(self, sheet: str, name: str, image: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        img = openpyxl.drawing.image.Image(fr'{ self.screenshots }/{ image }')
        img.anchor(ws.cell(self.get_image(sheet, name)))
        ws.add_image(img)
        wb.save(self.path)


@dataclass
class TestsPage:

    config: str = ConfigReader()

    def read(self) -> list[str]:
        path = self.config.read('path', 'tests_page')
        with open(path) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            for each_row in csv_reader:
                return each_row


config = ConfigReader()


def test() -> None:
    csvReader = csv.DictReader(open(config.read('path', 'tests_page')))
    for row in csvReader:
        print(row)


if __name__ == '__main__':
    test()
